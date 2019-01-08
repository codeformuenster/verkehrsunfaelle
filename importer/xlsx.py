from kinto_utils import create_accident_raw
from tqdm import tqdm

from openpyxl import load_workbook


def import_xlsx(file_path, file_meta, position):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[file_meta['sheet_name']]

    for row in tqdm(ws.iter_rows(min_row=file_meta['first_data_row']),
                    desc=file_meta['source_file'],
                    position=position,
                    unit='row',
                    dynamic_ncols=True):
        row_number = row[0].row
        raw_accident = {
            # column_name: str(ws.cell(row=row_number,
            #                          column=column_number).value)
            column_name: str(row[column_number - 1].value)
            for (column_name, column_number)
            in file_meta['columns_mapping'].items()
            if column_number
        }

        raw_accident['source_row_number'] = row_number

        for key in ['source_file', 'import_timestamp', 'source_file_hash']:
            raw_accident[key] = file_meta[key]

        create_accident_raw(raw_accident)
