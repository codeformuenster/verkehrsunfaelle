from kinto_utils import create_accident_raw, get_schema
from tqdm import tqdm
from datetime import datetime, time
from utils import accident_is_valid
import sys

from openpyxl import load_workbook

schema = get_schema('accidents_raw')


def transform_value(value, key):
    if value is None:
        return value

    if isinstance(value, datetime):
        return str(value)[0:10]

    if isinstance(value, time):
        return str(value)

    # find in schema
    schema_property_type = schema[key]['type']
    if schema_property_type == 'integer' and not isinstance(value, int):
        try:
            return int(value)
        except ValueError:
            return -1
    if schema_property_type == 'string' and not isinstance(value, str):
        if key == 'date':
            print(f'date {value} not of type datetime')
        if key == 'time_of_day':
            print(f'time_of_day {value} not of type time')
        return str(value)

    return value


def import_xlsx(file_path, file_meta, position):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[file_meta['sheet_name']]

    # for row in tqdm(ws.iter_rows(min_row=file_meta['first_data_row']),
    #                 desc=file_meta['source_file'],
    #                 position=position,
    #                 unit='row',
    #                 dynamic_ncols=True):
    for row in ws.iter_rows(min_row=file_meta['first_data_row']):
        row_number = row[0].row
        raw_accident = {
            column_name: transform_value(
                row[column_number - 1].value, column_name)
            for (column_name, column_number)
            in file_meta['columns_mapping'].items()
            if column_number and row[column_number - 1].value is not None
        }

        raw_accident['source_row_number'] = row_number

        for key in ['source_file', 'import_timestamp', 'source_file_hash']:
            raw_accident[key] = file_meta[key]

        if accident_is_valid(raw_accident) == True:
            create_accident_raw(raw_accident)
